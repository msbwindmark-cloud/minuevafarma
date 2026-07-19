from django.shortcuts import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from datetime import timedelta
from django.utils import timezone
import csv, io

from farmacia.models import Producto, Venta, DetalleVenta, Cliente, MovimientoStock, Caja, Categoria, Proveedor
from django.db.models import F, Q


def _productos_qs(alerta=''):
    hoy = timezone.now().date()
    limite = hoy + timedelta(days=30)
    qs = Producto.objects.select_related('categoria', 'proveedor')
    if alerta == 'caducado':
        qs = qs.filter(caducidad__lt=hoy)
    elif alerta == 'por_caducar':
        qs = qs.filter(caducidad__gte=hoy, caducidad__lte=limite)
    elif alerta == 'stock_bajo':
        qs = qs.filter(stock_actual__lte=F('stock_minimo'))
    return qs.all()


@login_required
def export_productos(request, fmt):
    alerta = request.GET.get('alerta', '')
    productos = _productos_qs(alerta)
    titulo = 'Productos'
    if alerta == 'caducado':
        titulo = 'Productos caducados'
    elif alerta == 'por_caducar':
        titulo = 'Productos por caducar'
    elif alerta == 'stock_bajo':
        titulo = 'Productos stock bajo'
    cols = ['Codigo', 'Nombre', 'Categoria', 'Ubicacion', 'Stock', 'Min', 'P.Compra', 'P.Venta', 'Caducidad']
    filas = [[p.codigo, p.nombre, p.categoria.nombre if p.categoria else '', p.ubicacion or '',
              p.stock_actual, p.stock_minimo, float(p.precio_compra), float(p.precio_venta),
              p.caducidad.strftime('%d/%m/%Y') if p.caducidad else ''] for p in productos]
    if fmt == 'excel':
        return _excel(titulo, cols, filas)
    return _pdf(titulo, cols, filas)


@login_required
def export_ventas(request, fmt):
    ventas = Venta.objects.filter(anulada=False).select_related('empleado').order_by('-fecha')[:500]
    cols = ['Codigo', 'Fecha', 'Cliente', 'Metodo', 'Total', 'Empleado']
    filas = [[v.codigo, v.fecha.strftime('%d/%m/%Y %H:%M'), v.cliente_nombre or '',
              v.get_metodo_pago_display(), float(v.total), str(v.empleado) if v.empleado else ''] for v in ventas]
    if fmt == 'excel':
        return _excel('Ventas', cols, filas)
    return _pdf('Ventas', cols, filas)


@login_required
def export_clientes(request, fmt):
    clientes = Cliente.objects.all()
    cols = ['Nombre', 'Telefono', 'Email', 'Puntos', 'Alergias']
    filas = [[c.nombre, c.telefono or '', c.email or '', c.puntos, c.alergias or ''] for c in clientes]
    if fmt == 'excel':
        return _excel('Clientes', cols, filas)
    return _pdf('Clientes', cols, filas)


@login_required
def export_movimientos(request, fmt):
    movs = MovimientoStock.objects.select_related('producto', 'usuario').order_by('-fecha')[:500]
    cols = ['Fecha', 'Producto', 'Tipo', 'Cantidad', 'Stock resultante', 'Motivo', 'Usuario']
    filas = [[m.fecha.strftime('%d/%m/%Y %H:%M'), m.producto.nombre if m.producto else '',
              m.get_tipo_display(), m.cantidad, m.stock_resultante, m.motivo or '',
              str(m.usuario) if m.usuario else ''] for m in movs]
    if fmt == 'excel':
        return _excel('Movimientos', cols, filas)
    return _pdf('Movimientos', cols, filas)


@login_required
def export_caja(request, fmt):
    cajas = Caja.objects.all()[:200]
    cols = ['Fecha', 'Tipo', 'Fondo', 'Efectivo', 'Tarjeta', 'Total', 'Usuario']
    filas = [[c.fecha.strftime('%d/%m/%Y %H:%M'), c.get_tipo_display(), float(c.fondo_inicial),
              float(c.efectivo), float(c.tarjeta), float(c.total),
              str(c.usuario) if c.usuario else ''] for c in cajas]
    if fmt == 'excel':
        return _excel('Caja', cols, filas)
    return _pdf('Caja', cols, filas)


def _excel(titulo, cols, filas):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:30]
    head_fill = PatternFill('solid', fgColor='0d6efd')
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = head_fill
    for i, fila in enumerate(filas, 2):
        for j, val in enumerate(fila, 1):
            ws.cell(row=i, column=j, value=val)
    for j in range(1, len(cols) + 1):
        ws.column_dimensions[chr(64 + j) if j <= 26 else 'A'].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % titulo.replace(' ', '_')
    return resp


def _pdf(titulo, cols, filas):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm,
                            leftMargin=12 * mm, rightMargin=12 * mm)
    styles = getSampleStyleSheet()
    elems = [Paragraph('<b>MiNuevaFarma - %s</b>' % titulo, styles['Title']), Spacer(1, 8)]
    data = [[Paragraph('<b>%s</b>' % c, styles['Normal']) for c in cols]] + [
        [Paragraph(str(v), styles['Normal']) for v in f] for f in filas]
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f6fb')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="%s.pdf"' % titulo.replace(' ', '_')
    return resp


@login_required
def export_categorias(request, fmt):
    cats = Categoria.objects.all()
    cols = ['Nombre', 'Descripcion']
    filas = [[c.nombre, c.descripcion or ''] for c in cats]
    if fmt == 'excel':
        return _excel('Categorias', cols, filas)
    return _pdf('Categorias', cols, filas)


@login_required
def export_proveedores(request, fmt):
    provs = Proveedor.objects.all()
    cols = ['Nombre', 'CIF', 'Telefono', 'Email', 'Direccion']
    filas = [[p.nombre, p.cif or '', p.telefono or '', p.email or '', p.direccion or ''] for p in provs]
    if fmt == 'excel':
        return _excel('Proveedores', cols, filas)
    return _pdf('Proveedores', cols, filas)


@login_required
def export_alertas(request, fmt):
    hoy = timezone.now().date()
    limite = hoy + timedelta(days=30)
    caducados = list(Producto.objects.filter(caducidad__lt=hoy, activo=True))
    por_caducar = list(Producto.objects.filter(caducidad__gte=hoy, caducidad__lte=limite, activo=True))
    stock_bajo = list(Producto.objects.filter(stock_actual__lte=F('stock_minimo'), activo=True))
    cols = ['Tipo', 'Producto', 'Stock', 'Min', 'Caducidad']
    filas = []
    for p in caducados:
        filas.append(['CADUCADO', p.nombre, p.stock_actual, p.stock_minimo, p.caducidad.strftime('%d/%m/%Y') if p.caducidad else ''])
    for p in por_caducar:
        filas.append(['POR CADUCAR', p.nombre, p.stock_actual, p.stock_minimo, p.caducidad.strftime('%d/%m/%Y') if p.caducidad else ''])
    for p in stock_bajo:
        filas.append(['STOCK BAJO', p.nombre, p.stock_actual, p.stock_minimo, p.caducidad.strftime('%d/%m/%Y') if p.caducidad else ''])
    if fmt == 'excel':
        return _excel('Alertas', cols, filas)
    return _pdf('Alertas', cols, filas)

